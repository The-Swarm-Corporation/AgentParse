import os
from PyPDF2 import PdfReader
import openpyxl
from loguru import logger
from swarms.models.tiktoken_wrapper import TikTokenizer
from typing import List


def file_to_string(file_path: str) -> str:
    """
    Convert various file types to string, auto-detecting the file extension.
    Supported types: .txt, .csv, .pdf, .docx, .xlsx, .json

    Args:
    file_path (str): Path to the file

    Returns:
    str: Content of the file as a string
    """
    _, file_extension = os.path.splitext(file_path)

    try:
        if file_extension == ".txt":
            with open(file_path, "r", encoding="utf-8") as file:
                return file.read()

        elif file_extension == ".csv":
            with open(file_path, "r", encoding="utf-8") as file:
                return file.read()

        elif file_extension == ".pdf":
            text = ""
            with open(file_path, "rb") as file:
                pdf_reader = PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text

        elif file_extension == ".xlsx":
            wb = openpyxl.load_workbook(file_path)
            text = ""
            for sheet in wb.sheetnames:
                text += f"Sheet: {sheet}\n"
                for row in wb[sheet].iter_rows(values_only=True):
                    text += ",".join(str(cell) for cell in row) + "\n"
            return text

        elif file_extension == ".json":
            with open(file_path, "r", encoding="utf-8") as file:
                return file.read()

        else:
            raise ValueError(
                f"Unsupported file type: {file_extension}"
            )

    except Exception as e:
        logger.error(f"Error processing file {file_path}: {str(e)}")
        raise


def chunk_text_dynamic(
    text: str, limit_tokens: int = 10000
) -> List[str]:
    """
    Chunk text into smaller chunks based on the token limit, ensuring words are not cut off.

    Args:
    text (str): The input text to be chunked
    limit_tokens (int): The approximate number of tokens per chunk (default: 10000)

    Returns:
    List[str]: A list of text chunks
    """
    tokenizer = TikTokenizer()
    words = text.split()
    chunks = []
    current_chunk = []
    current_token_count = 0

    for word in words:
        word_tokens = tokenizer.count_tokens(word)
        if (
            current_token_count + word_tokens > limit_tokens
            and current_chunk
        ):
            chunks.append(" ".join(current_chunk))
            current_chunk = []
            current_token_count = 0

        current_chunk.append(word)
        current_token_count += word_tokens

    if current_chunk:
        chunks.append(" ".join(current_chunk))

    return chunks
